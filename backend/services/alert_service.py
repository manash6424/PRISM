"""
Export service for generating CSV, Excel, and PDF exports.
"""
import os
import csv
import json
import logging
from typing import Optional, List
from datetime import datetime

from ..models.database import QueryResponse, ExportFormat, ExportRequest
from ..config import get_settings

logger = logging.getLogger(__name__)


class ExportService:
    """
    Handles export of query results to various formats.
    Supports CSV, Excel, PDF, and JSON exports.
    """

    def __init__(self):
        self.settings = get_settings()
        self._ensure_export_dir()

    def _ensure_export_dir(self):
        export_dir = self.settings.export.export_dir
        if not os.path.exists(export_dir):
            os.makedirs(export_dir, exist_ok=True)

    async def export(self, request: ExportRequest, query_response: QueryResponse) -> str:
        if not query_response.success:
            raise ValueError("Cannot export failed query")

        base_filename = request.filename or f"export_{query_response.query_id[:8]}"

        if request.format == ExportFormat.CSV:
            return await self._export_csv(base_filename, query_response)
        elif request.format == ExportFormat.EXCEL:
            return await self._export_excel(base_filename, query_response, request)
        elif request.format == ExportFormat.PDF:
            return await self._export_pdf(base_filename, query_response, request)
        elif request.format == ExportFormat.JSON:
            return await self._export_json(base_filename, query_response)
        else:
            raise ValueError(f"Unsupported export format: {request.format}")

    async def _export_csv(self, filename: str, query_response: QueryResponse) -> str:
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.csv")

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if query_response.columns:
                writer.writerow(query_response.columns)
            for row in query_response.results:
                writer.writerow([row.get(col, "") for col in query_response.columns])

        logger.info(f"Exported to CSV: {filepath}")
        return filepath

    async def _export_excel(self, filename: str, query_response: QueryResponse, request: ExportRequest) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")

        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, column in enumerate(query_response.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, row in enumerate(query_response.results, 2):
            for col_idx, column in enumerate(query_response.columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(column, ""))

        for col_idx in range(1, len(query_response.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(query_response.columns[col_idx - 1])
            for row in query_response.results:
                value = str(row.get(query_response.columns[col_idx - 1], ""))
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filepath)
        logger.info(f"Exported to Excel: {filepath}")
        return filepath

    async def _export_pdf(self, filename: str, query_response: QueryResponse, request: ExportRequest) -> str:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            raise ImportError("reportlab required: pip install reportlab")

        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.pdf")

        # FIX: max_rows logic moved outside html, with warning log
        max_rows = getattr(request, 'max_rows', 1000)
        if query_response.row_count > max_rows:
            logger.warning(f"PDF export truncated to {max_rows} rows (total: {query_response.row_count})")
        rows_to_show = query_response.results[:max_rows]

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_text = request.title or "Query Results"
        elements.append(Paragraph(title_text, styles['Title']))
        elements.append(Spacer(1, 12))

        # Meta info
        meta = (
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC | "
            f"Query: {query_response.natural_language[:100]}... | "
            f"Rows: {query_response.row_count} | "
            f"Execution Time: {query_response.execution_time_ms:.2f}ms"
        )
        elements.append(Paragraph(meta, styles['Normal']))
        elements.append(Spacer(1, 12))

        # Build table data — header + rows
        table_data = [query_response.columns]
        for row in rows_to_show:
            table_data.append([str(row.get(col, "")) for col in query_response.columns])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#4472C4')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0, 0), (-1, 0),  10),
            ('ALIGN',          (0, 0), (-1, 0),  'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE',       (0, 1), (-1, -1), 8),
            ('PADDING',        (0, 0), (-1, -1), 6),
            ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)

        # Optional description
        if request.description:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(request.description, styles['Italic']))

        doc.build(elements)
        logger.info(f"Exported to PDF: {filepath}")
        return filepath

    async def _export_json(self, filename: str, query_response: QueryResponse) -> str:
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.json")

        export_data = {
            "query": {
                "id": query_response.query_id,
                "natural_language": query_response.natural_language,
                "generated_sql": query_response.generated_sql,
                "explanation": query_response.explanation,
                "timestamp": query_response.timestamp.isoformat(),
            },
            "results": {
                "columns": query_response.columns,
                "rows": query_response.results,
                "row_count": query_response.row_count,
            },
            "metadata": {
                "execution_time_ms": query_response.execution_time_ms,
                "exported_at": datetime.utcnow().isoformat(),
            }
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str, ensure_ascii=False)

        logger.info(f"Exported to JSON: {filepath}")
        return filepath

    def get_export_formats(self) -> List[str]:
        return [f.value for f in ExportFormat]


# Global export service
export_service = ExportService()