"""
Export service for generating CSV, Excel, and PDF exports.
"""
import os
import csv
import json
import logging
from typing import Optional, List
from datetime import datetime
 
from ..models.database import QueryResponse, ExportFormat, ExportRequest
from ..config import get_settings
 
logger = logging.getLogger(__name__)
 
 
class ExportService:
    def __init__(self):
        self.settings = get_settings()
        self._ensure_export_dir()
 
    def _ensure_export_dir(self):
        export_dir = self.settings.export.export_dir
        if not os.path.exists(export_dir):
            os.makedirs(export_dir, exist_ok=True)
 
    async def export(self, request: ExportRequest, query_response: QueryResponse) -> str:
        if not query_response.success:
            raise ValueError("Cannot export failed query")
 
        base_filename = request.filename or f"export_{query_response.query_id[:8]}"
 
        if request.format == ExportFormat.CSV:
            return await self._export_csv(base_filename, query_response)
        elif request.format == ExportFormat.EXCEL:
            return await self._export_excel(base_filename, query_response, request)
        elif request.format == ExportFormat.PDF:
            return await self._export_pdf(base_filename, query_response, request)
        elif request.format == ExportFormat.JSON:
            return await self._export_json(base_filename, query_response)
        else:
            raise ValueError(f"Unsupported export format: {request.format}")
 
    async def _export_csv(self, filename: str, query_response: QueryResponse) -> str:
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.csv")
 
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if query_response.columns:
                writer.writerow(query_response.columns)
            for row in query_response.results:
                writer.writerow([row.get(col, "") for col in query_response.columns])
 
        logger.info(f"Exported to CSV: {filepath}")
        return filepath
 
    async def _export_excel(self, filename: str, query_response: QueryResponse, request: ExportRequest) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")
 
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.xlsx")
 
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"
 
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
 
        for col_idx, column in enumerate(query_response.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
 
        for row_idx, row in enumerate(query_response.results, 2):
            for col_idx, column in enumerate(query_response.columns, 1):
                val = row.get(column, "")
                if val is not None:
                    val = str(val)
                ws.cell(row=row_idx, column=col_idx, value=val)
 
        for col_idx in range(1, len(query_response.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(query_response.columns[col_idx - 1])
            for row in query_response.results:
                value = str(row.get(query_response.columns[col_idx - 1], ""))
                max_length = max(max_length, len(value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
 
        wb.save(filepath)
        logger.info(f"Exported to Excel: {filepath}")
        return filepath
 
    async def _export_pdf(self, filename: str, query_response: QueryResponse, request: ExportRequest) -> str:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise ImportError("reportlab required: pip install reportlab")
 
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.pdf")
 
        pagesize = landscape(A4) if len(query_response.columns) > 4 else A4
        doc = SimpleDocTemplate(
            filepath,
            pagesize=pagesize,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1*cm
        )
 
        styles = getSampleStyleSheet()
        elements = []
 
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=10
        )
        elements.append(Paragraph(request.title or "Query Results", title_style))
 
        meta_style = ParagraphStyle(
            'Meta',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            spaceAfter=15
        )
        generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        elements.append(Paragraph(
            f"Generated: {generated_at} | Rows: {query_response.row_count} | Time: {query_response.execution_time_ms:.0f}ms",
            meta_style
        ))
 
        if query_response.natural_language:
            elements.append(Paragraph(f"Query: {query_response.natural_language[:200]}", meta_style))
 
        elements.append(Spacer(1, 0.3*cm))
 
        table_data = [query_response.columns]
        for row in query_response.results[:1000]:
            table_row = []
            for col in query_response.columns:
                val = str(row.get(col, ""))
                if len(val) > 50:
                    val = val[:47] + "..."
                table_row.append(val)
            table_data.append(table_row)
 
        available_width = pagesize[0] - 2*cm
        col_width = available_width / len(query_response.columns)
        col_widths = [col_width] * len(query_response.columns)
 
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EEF2FF')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#2E5EA8')),
        ]))
 
        elements.append(table)
        doc.build(elements)
        logger.info(f"Exported to PDF: {filepath}")
        return filepath
 
    async def _export_json(self, filename: str, query_response: QueryResponse) -> str:
        filepath = os.path.join(self.settings.export.export_dir, f"{filename}.json")
 
        export_data = {
            "query": {
                "id": query_response.query_id,
                "natural_language": query_response.natural_language,
                "generated_sql": query_response.generated_sql,
                "explanation": query_response.explanation,
                "timestamp": query_response.timestamp.isoformat(),
            },
            "results": {
                "columns": query_response.columns,
                "rows": query_response.results,
                "row_count": query_response.row_count,
            },
            "metadata": {
                "execution_time_ms": query_response.execution_time_ms,
                "exported_at": datetime.utcnow().isoformat(),
            }
        }
 
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str, ensure_ascii=False)
 
        logger.info(f"Exported to JSON: {filepath}")
        return filepath
 
    def get_export_formats(self) -> List[str]:
        return [f.value for f in ExportFormat]
 
 
# Global export service
export_service = ExportService()